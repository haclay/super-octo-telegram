# python alibaba web crawler 1.0

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from lxml import etree
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
import requests
import threading
import queue
import time
import random
import re
import sys

# 获取客户数据
search = input("请输入查询内容: ")
page_nums = int(input("总共需要查询多少页数据: "))
excel_name_ = input("请命名Excel表(可根据产品名称): ")
thread_num = round((page_nums * 50)/40)
# 调整间隔时间并开始计时
start_time = time.time()
sleep_time = random.randint(2, 5) + random.random()
# 选择谷歌浏览器
service = Service(executable_path=r'.\chromedriver.exe')
driver = webdriver.Chrome(service=service)
# 获取查询页面源码
link = "https://www.alibaba.com/"
driver.get(link)
time.sleep(sleep_time)
element_search_input = driver.find_element(By.CLASS_NAME, 'search-bar-input')
time.sleep(sleep_time)
element_search_input.send_keys(search, Keys.ENTER)
time.sleep(sleep_time)
# 获取产品网页链接
product_name_list = []  # 产品名称
product_href_list = []  # 产品网页链接
page_num = 1
error_time = 0
while page_num <= page_nums:
    # 数据解析
    page_text = driver.page_source
    html = etree.HTML(page_text)
    product_sections = html.xpath('//div[@class="organic-list app-organic-search__list"]/div')
    print(f"第{page_num}张网页产品网页链接正在获取")
    for index, product in enumerate(product_sections):
        if index == 0 and page_num == 1:
            try:
                href = product.xpath('./div/div/div/div/div[2]/div[1]/div[1]/a/@href')[0]
                href_ = "https:" + href
                product_href_list.append(href_)
            except IndexError:
                try:
                    href = product.xpath('./div/div[2]/div[1]/h2/a/@href')[0]
                    href_ = "https:" + href
                    product_href_list.append(href_)
                except IndexError:
                    try:
                        href = product.xpath('./div/div/div/div/div[2]/div[1]/a/@href')[0]
                        href_ = "https:" + href
                        product_href_list.append(href_)
                    except IndexError:
                        print("当前网页第一个产品链接无法获取, 请重新运行程序")
        else:
            try:
                href = product.xpath('./div/div[2]/div[1]/h2/a/@href')[0]
                href_ = "https:" + href
                product_href_list.append(href_)
            except IndexError:
                error_time += 1
    print(f"第{page_num}张网页产品网页链接获取完成")
    print("当前总产品数量: ", len(product_href_list))
    print("当前未解析的网页链接总数量:", error_time)
    # 进行下一页操作
    page_num += 1
    if page_num > page_nums:
        print("产品数据爬取结束")
        if len(product_href_list) % 48 == 0:
            print("产品网页链接获取成功\n\n")
            driver.quit()
        else:
            print("某个产品网页链接获取出现异常,未获取的网页链接数量:", error_time, "\n\n")
            driver.quit()
        break
    else:
        element_go_to_page = driver.find_element(By.CLASS_NAME, 'seb-pagination__goto-input')
        element_go_to_page_trigger = driver.find_element(By.CLASS_NAME, 'seb-pagination__goto-trigger')
        element_go_to_page.send_keys(page_num)
        time.sleep(sleep_time)
        element_go_to_page_trigger.click()
        time.sleep(5)
# 获取单个产品页面数据并解析公司网站链接
# 创建多线程
print("正在获取供应商页面链接")


def get_user_agent():
    ua = [
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/114.0.5735.110 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/114.0.5735.134 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/110.0',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/111.0',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/112.0',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/114.0.5735.110 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/114.0.5735.134 Safari/537.36',
    ]
    user_agent = random.choice(ua)
    return user_agent


class My_thread(threading.Thread):
    def __init__(self, name, que):
        threading.Thread.__init__(self)
        self.name = name
        self.que = que
        self.thread_results = []

    def run(self):
        print("Starting:", self.name)
        while True:
            try:
                sp = self.que.get(timeout=2)
                results = get_single_product_page(self.name, self.que, sp)
                self.thread_results.append(results)
                time.sleep(random.randint(2, 3) + random.random())
            except queue.Empty:
                break
        print("Existing:", self.name)


def get_single_product_page(thread_name, q, single_product):
    # 定义线程执行的任务 get_requests_html(), parses_data()
    try:
        headers = {
            'User-Agent': get_user_agent(),
            'Host': 'www.alibaba.com',
            'Referer': f'https://www.alibaba.com/trade/search?spm=a2700.product_home_newuser.'
                       f'home_new_user_first_screen_fy23_pc_search_bar.keydown__Enter&tab=all&'
                       f'searchText={search}',
        }
        single_page_text = requests.get(single_product, headers=headers).text
        time.sleep(random.random())
        page_html = etree.HTML(single_page_text)
        company_href = parses_data(thread_name, q, single_product, page_html)
        return company_href
    except requests.exceptions.InvalidURL as e:
        sp_without_protocol = single_product[len('https:'):]
        page_html = get_request_html(sp_without_protocol)
        company_href = parses_data(thread_name, q, sp_without_protocol, page_html)
        return company_href
    except requests.exceptions.RequestException as e:
        print("请求异常:", e)
    except Exception as e:
        print("其他异常:", e)


def get_request_html(url):
    headers = {
        'User-Agent': get_user_agent(),
    }
    single_page_text = requests.get(url, headers=headers).text
    time.sleep(random.random())
    page_html = etree.HTML(single_page_text)
    return page_html


def parses_data(thread_name, q, single_product, page_html):
    try:
        company_href = page_html.xpath('//div[@class="company-head"]/div[2]/a/@href')[0]
        print(thread_name, q.qsize())
        return company_href
    except IndexError:
        try:
            company_href = page_html.xpath('//div[@class="company-name-container"]/a/@href')[0]
            print(thread_name, q.qsize())
            return company_href
        except IndexError:
            print(thread_name, q.qsize(), "无法获取公司网站链接:", single_product)


def create_thread_queue(links):
    thread_list = [f'thread_{t+1}' for t in range(thread_num)]
    workqueue = queue.Queue(len(product_href_list[:]))
    # 填充队列
    for link in links:
        workqueue.put(link)
    # 创建新线程
    threads = []
    for thread_name in thread_list:
        my_thread = My_thread(thread_name, workqueue)
        my_thread.start()
        threads.append(my_thread)
    for thread_ in threads:
        thread_.join()
    # 合并线程的结果
    company_hrefs = []
    for thread_ in threads:
        company_hrefs.extend(thread_.thread_results)
    return company_hrefs


company_hrefs_ = create_thread_queue(product_href_list[:])
none_num = company_hrefs_.count(None)
len_company_hrefs_ = len(company_hrefs_) - none_num
if none_num == 0:
    print(f"产品网页链接爬取成功, 获取总产品链接数量:{len_company_hrefs_}\n\n")
elif none_num <= 50:
    print("无法获取部分公司网站链接, 成功爬取公司网站链接总数量:", len_company_hrefs_, "\n\n")
    company_hrefs_ = list(filter(None, company_hrefs_))
else:
    print("无法获取部分公司网站链接, 成功爬取公司网站链接总数量:", len_company_hrefs_, "\n\n")
    print("程序已退出")
    sys.exit(-2)
# # 获取公司基本信息页面的链接
print("正在获取供应商页面contacts链接")


class My_Thread(threading.Thread):
    def __init__(self, name, que):
        threading.Thread.__init__(self)
        self.name = name
        self.que = que
        self.thread_results = []

    def run(self):
        print("Starting:", self.name)
        while True:
            try:
                cl = self.que.get(timeout=2)
                results = get_company_page(self.name, self.que, cl)
                self.thread_results.append(results)
                time.sleep(random.randint(2, 3) + random.random())
            except queue.Empty:
                break
        print("Existing:", self.name)


def get_company_page(thread_name, q, company_link):
    # 定义线程执行的任务
    try:
        company_html = get_request_html(company_link)
        try:
            company_contacts = company_html.xpath('//ul[@class="navigation-list"]/li[4]/a/@href')[0]
            # 提取网站特定头部
            pattern = r'https://(.*?)/'
            matches = re.match(pattern, company_link)[0]
            link_header = matches.rstrip('/')
            company_contacts = link_header + company_contacts
            print(thread_name, q.qsize())
            return company_contacts
        except IndexError:
            print(thread_name, q.qsize(), "无法获取公司contacts链接:", company_link)
    except Exception as e:
        print(thread_name, q.qsize(), e)


def create_thread_queue(links):
    thread_list = [f'thread_{t+1}' for t in range(thread_num)]
    workqueue = queue.Queue(len(company_hrefs_[:]))
    # 填充队列
    for link_ in links:
        workqueue.put(link_)
    # 创建新线程
    threads = []
    for thread_name in thread_list:
        my_thread = My_Thread(thread_name, workqueue)
        my_thread.start()
        threads.append(my_thread)
    for thread_ in threads:
        thread_.join()
    # 合并线程的结果
    company_contacts_link = []
    for thread_ in threads:
        company_contacts_link.extend(thread_.thread_results)
    return company_contacts_link


company_contacts_link_ = create_thread_queue(company_hrefs_[:])
none_num_ = company_contacts_link_.count(None)
len_company_contacts_link = len(company_contacts_link_) - none_num_
if none_num_ == 0:
    print("公司contacts链接爬取成功, 获取总链接数量:", len_company_contacts_link, "\n\n")
if none_num_ <= 10:
    print("无法获取部分公司contacts链接, 获取总链接数量:", len_company_contacts_link, "\n\n")
    company_contacts_link_ = list(filter(None, company_contacts_link_))
else:
    print("无法获取部分公司contacts链接, 获取链接数量:", len_company_contacts_link)
    print("退出程序")
    sys.exit(-3)
# 获取公司基本信息
print("正在获取供应商基本信息")
company_name_list = []
operational_address_list = []
website_list = []
website_on_alibaba_list = []
for detail_link in company_contacts_link_:
    not_parse_num = len(company_contacts_link_) - len(company_name_list)
    print("正在解析网址:", detail_link)
    print("未解析的网址数量:", not_parse_num)
    detail_html = get_request_html(detail_link)
    try:
        company_name = detail_html.xpath('//table[@class="contact-table"]/tr[1]/td/text()')[0]
        company_name_list.append(company_name)
    except IndexError:
        company_name_list.append(f'未解析出公司名称, 供应商contacts网址: {detail_link}')
    try:
        operational_address = detail_html.xpath('//table[@class="contact-table"]/tr[2]/td/text()')[0]
        operational_address_list.append(operational_address)
    except IndexError:
        operational_address_list.append(f'未解析出公司地址, 供应商contacts网址: {detail_link}')
    try:
        website = detail_html.xpath('//table[@class="contact-table"]/tr[3]/td/div/text()')[0]
        website_list.append(website)
    except IndexError:
        website_list.append('未解析出公司网站')
        try:
            website_on_alibaba = detail_html.xpath('//table[@class="contact-table"]/tr[3]/td/a/@href')[0]
            website_on_alibaba_ = 'https:' + website_on_alibaba
            website_on_alibaba_list.append(website_on_alibaba_)
        except IndexError:
            website_on_alibaba_list.append(f'未解析出公司alibaba网站, 供应商contacts网址: {detail_link}')
        continue
    try:
        website_on_alibaba = detail_html.xpath('//table[@class="contact-table"]/tr[4]/td/a/@href')[0]
        website_on_alibaba_ = 'https:' + website_on_alibaba
        website_on_alibaba_list.append(website_on_alibaba_)
    except IndexError:
        website_on_alibaba_list.append(f'未解析出公司alibaba网站, 供应商contacts网址: {detail_link}')
    time.sleep(random.random())
site_num = len(website_list) - website_list.count('未解析出公司网站')
print("数据解析全部完成")
print("获取公司名称的总数量:", len(company_name_list))
print("获取公司地址的总数量:", len(operational_address_list))
print("获取公司网址的总数量:", site_num)
print("获取公司alibaba网址总数量:", len(website_on_alibaba_list), "\n\n")
end_time = time.time()
print("本次爬虫结束")
print("本次爬虫耗时:", end_time-start_time)
# 数据存储
print("正在进行数据存储")


def store_data(cnl, oal, wl, wal):
    num_list = [i+1 for i in range(len(cnl))]
    data_dict = {
        'Number': num_list,
        'Company Name': cnl,
        'Operational Address': oal,
        'Website': wl,
        'Website on alibaba.com': wal,
    }
    dataframe = pd.DataFrame(data_dict)
    filename = f"{excel_name_}.xlsx"
    dataframe.to_excel(filename, index=False)
    print("下载成功")
# 自适应列宽


def get_num_columns_dict():
    num_str_dict = {}
    a_z = [chr(a) for a in range(ord('A'), ord('Z') + 1)]
    aa_az = ['A' + chr(a) for a in range(ord('A'), ord('Z') + 1)]
    a_az = a_z + aa_az
    for i in a_az:
        num_str_dict[a_az.index(i) + 1] = i
    return num_str_dict


def style_excel(excel_name: str, sheet_name: str):
    wb = openpyxl.load_workbook(excel_name)
    sheet = wb[sheet_name]
    max_column = sheet.max_column
    max_row = sheet.max_row
    max_column_dict = {}
    num_str_dict = get_num_columns_dict()
    for i in range(1, max_column + 1):
        for j in range(1, max_row + 1):
            column = 0
            sheet_value = sheet.cell(row=j, column=i).value
            sheet_value_list = [k for k in str(sheet_value)]
            for v in sheet_value_list:
                if v.isdigit() is True or v.isalpha() is True:
                    column += 1.1
                else:
                    column += 2.2
            try:
                if column > max_column_dict[i]:
                    max_column_dict[i] = column
            except Exception as e:
                print(e)
                max_column_dict[i] = column
    for key, value in max_column_dict.items():
        sheet.column_dimensions[num_str_dict[key]].width = value
    # 保存
    wb.save(excel_name)
    print("excel表格调整完成")


def create_hyperlink(excel_name, sheet_name):
    wb = load_workbook(f'{excel_name_}.xlsx')
    ws = wb['Sheet1']
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=5)
        url = cell.value
        cell.font = Font(underline='single', color='0000FF')
        cell.hyperlink = url
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=4)
        url = cell.value
        cell.font = Font(underline='single')
        cell.hyperlink = url
    wb.save(f"{excel_name_}.xlsx")
    print("超链接已创建完成")


if len(company_name_list) == len(operational_address_list) == \
        len(website_list) == len(website_on_alibaba_list):
    store_data(company_name_list, operational_address_list, website_list, website_on_alibaba_list)
    style_excel(f'{excel_name_}.xlsx', 'Sheet1')
    create_hyperlink(f'{excel_name_}.xlsx', 'Sheet1')
    print("程序执行完毕, Excel表格已经创建到当前文件中\n")
else:
    print("列表长度存在问题, 未能生成excel表")

time.sleep(10)


